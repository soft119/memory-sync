#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
import sys
import os
from openpyxl import load_workbook

def add_7day_card(file_path):
    """在cfg_item.xls中添加7天点卡物品"""
    try:
        print(f"正在读取Excel文件: {file_path}")
        
        # 备份原文件
        backup_path = file_path.replace('.xls', '_backup.xls')
        import shutil
        shutil.copy2(file_path, backup_path)
        print(f"已创建备份: {backup_path}")
        
        # 读取Excel文件（保持原始结构）
        wb = load_workbook(file_path)
        ws = wb.active
        
        print(f"工作表: {ws.title}, 最大行: {ws.max_row}, 最大列: {ws.max_column}")
        
        # 找到数据开始行（第4行开始是物品数据）
        data_start_row = 4  # 第1行：//;ver，第2行：中文说明，第3行：英文列名，第4行开始：数据
        
        # 找到最后一个物品的行
        last_item_row = ws.max_row
        print(f"最后一个物品在第 {last_item_row} 行")
        
        # 获取最后一个物品的Idx
        last_idx = ws.cell(row=last_item_row, column=1).value
        print(f"最后一个物品Idx: {last_idx}")
        
        # 新物品Idx = 最后Idx + 1
        new_idx = int(last_idx) + 1 if last_idx else 10345
        print(f"新物品Idx: {new_idx}")
        
        # 添加新行
        new_row = last_item_row + 1
        
        # 7天点卡的属性
        item_data = {
            1: new_idx,           # Idx
            2: "7天点卡",         # Name
            3: 2,                 # StdMode (计次物品)
            4: 0,                 # Shape (外观)
            5: 0,                 # Weight (重量)
            6: 100,               # Anicount (关键！对应[@StdModeFunc100])
            7: 0,                 # Source
            8: 0,                 # Reserved
            9: 266,               # Looks (幻境凭证图标ID)
            10: 1000,             # DuraMax (持久度=1000 → 使用次数=1)
            11: "",               # Attribute
            12: 0,                # Need
            13: 0,                # NeedLevel
            14: 0,                # Price (0=不可购买)
            15: 0,                # Color
            16: 0,                # OverLap
            17: 0,                # Suit
            18: "",               # Article
            19: "",               # Job
            20: "",               # effectParam
            21: "双击增加7天游戏时间",  # Desc (备注)
            22: "",               # Expand1
            23: "",               # HairShow
            24: "",               # auctionby
            25: ""                # Insurance
        }
        
        # 写入数据
        for col, value in item_data.items():
            ws.cell(row=new_row, column=col, value=value)
        
        print(f"在第 {new_row} 行添加了7天点卡")
        
        # 保存文件
        wb.save(file_path)
        print(f"文件已保存: {file_path}")
        
        print("\n" + "="*60)
        print("7天点卡添加完成！")
        print(f"物品Idx: {new_idx}")
        print("关键属性：")
        print(f"  Name: 7天点卡")
        print(f"  StdMode: 2 (计次物品)")
        print(f"  Anicount: 100 (对应[@StdModeFunc100]标签)")
        print(f"  Looks: 266 (幻境凭证图标)")
        print(f"  DuraMax: 1000 (持久度=1000 → 使用次数=1)")
        print(f"  Price: 0 (不可购买)")
        
        # 验证添加结果
        print("\n" + "="*60)
        print("验证添加结果：")
        
        # 重新读取验证
        df = pd.read_excel(file_path, header=2)
        new_item = df[df['Name'] == '7天点卡']
        if len(new_item) > 0:
            print("✅ 7天点卡已成功添加到数据库！")
            row = new_item.iloc[0]
            print(f"  物品ID: {row.get('Idx')}")
            print(f"  Anicount: {row.get('Anicount')}")
            print(f"  Looks: {row.get('Looks')}")
        else:
            print("❌ 添加失败，未找到7天点卡")
        
        return True
        
    except Exception as e:
        print(f"添加物品失败: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python add_7天点卡.py <cfg_item.xls路径>")
        sys.exit(1)
    
    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        print(f"文件不存在: {file_path}")
        sys.exit(1)
    
    success = add_7day_card(file_path)
    
    if success:
        print("\n✅ 7天点卡添加成功！")
        print("请重启游戏引擎或重新加载物品数据库")
    else:
        print("\n❌ 添加失败")